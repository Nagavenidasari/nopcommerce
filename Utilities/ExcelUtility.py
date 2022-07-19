import openpyxl
#import openpyexcel

def getRowscount(file,SheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[SheetName]
    return sheet.max_row
def getColumncount(file,SheetName):
    workbook = openpyxl.load_workbook(file)
    sheet =workbook[SheetName]
    return sheet.max_column
def readData(file,SheetName,rownum,columnnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[SheetName]
    return sheet.cell(row=rownum,column=columnnum).value
def writeDate(file,SheetName,rownum,columnnum,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[SheetName]
    sheet.cell(row=rownum,column=columnnum).value=data
    workbook.save(file)

